#!/usr/bin/env python3
import csv, json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter

OUT=Path('data/green_gregory')
CSV=OUT/'sam_balanced_intake_100.csv'
SUMMARY=OUT/'sam_balanced_intake_100_summary.json'
DEST=OUT/'GGG_Historical_Open_Solicitation_Intake_Test01.xlsx'

with CSV.open(encoding='utf-8',newline='') as f:
    rows=list(csv.DictReader(f))
summary=json.loads(SUMMARY.read_text(encoding='utf-8'))
wb=Workbook(); ws=wb.active; ws.title='Intake_100'
headers=list(rows[0])
ws.append(headers)
for r in rows: ws.append([r.get(h,'') for h in headers])
ws.freeze_panes='A2'; ws.auto_filter.ref=ws.dimensions
for c in ws[1]:
    c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='1F4E78'); c.alignment=Alignment(wrap_text=True,vertical='top')
for row in ws.iter_rows(min_row=2):
    for c in row: c.alignment=Alignment(wrap_text=True,vertical='top')
widths={1:12,2:34,3:24,4:45,5:24,6:12,7:18,8:28,9:28,10:30,11:45,12:65,13:12,14:12,15:26,16:20,17:24,18:32,19:35,20:35,21:28,22:25,23:20,24:24,25:45,26:32,27:16,28:20,29:20,30:28,31:24,32:30,33:30,34:24,35:32,36:36,37:26,38:22,39:24,40:22,41:18,42:18,43:22,44:34,45:24,46:18,47:70,48:18}
for i,w in widths.items(): ws.column_dimensions[get_column_letter(i)].width=w
ws.row_dimensions[1].height=42

funnel=wb.create_sheet('Funnel')
funnel.append(['Stage / Metric','Count','Interpretation'])
funnel_rows=[
 ['Official SAM full extract rows',78811,'Source population at retrieval'],
 ['Posted during 2023-2024',5739,'Primary historical period'],
 ['Actual Solicitation or Combined notice',1505,'Notice-type gate'],
 ['Response deadline no later than 2024-12-31',1024,'Closed historical solicitation universe'],
 ['Distinct solicitation families',summary['distinct_solicitation_families'],'Amendments and notice versions collapsed'],
 ['Balanced intake selected',100,'Selection independent of award outcome'],
 ['Pass Gates 1-6 preliminarily',summary['gate_counts'].get('Passes Gates 1-6 preliminarily; Gate 7 archive review pending',0),'Advance to archive/attachment review'],
 ['Killed at Gate 4 — GGG eligibility',summary['gate_counts'].get('Gate 4 — GGG eligibility',0),'Set-aside or eligibility mismatch'],
 ['Killed at Gate 5 — mandatory source',summary['gate_counts'].get('Gate 5 — mandatory source',0),'AbilityOne/FPI/mandatory-channel signal'],
 ['Killed at Gate 6 — product/service exclusion',summary['gate_counts'].get('Gate 6 — product/service exclusion',0),'Current kickoff exclusions'],
 ['Award outcomes already mapped',summary['mapped_outcomes'],'Outcome answer key recovered without driving selection'],
 ['Pass 1 outcome-conditioned review count',2,'Rejected sampling design'],
 ['Pass 1 outcome-conditioned rejection count',98,'88/100 DoD; amendment duplication and outcome bias'],
]
for r in funnel_rows: funnel.append(r)
funnel.freeze_panes='A2'
for c in funnel[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='1F4E78')
for col,w in {'A':48,'B':15,'C':70}.items(): funnel.column_dimensions[col].width=w
chart=BarChart(); chart.title='Historical Intake Funnel'; chart.y_axis.title='Records'; chart.x_axis.title='Stage'
data=Reference(funnel,min_col=2,min_row=2,max_row=11); cats=Reference(funnel,min_col=1,min_row=2,max_row=11)
chart.add_data(data,titles_from_data=False); chart.set_categories(cats); chart.height=9; chart.width=18
funnel.add_chart(chart,'E2')

tracks=wb.create_sheet('Supply_Tracks')
tracks.append(['Supply Track','Selected Count','Target','Shortage Before Fill'])
for tr,count in summary['supply_tracks'].items():
    tracks.append([tr,count,summary['target_quotas'].get(tr,''),summary['quota_shortage_before_fill'].get(tr,'')])
for c in tracks[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='1F4E78')
for col,w in {'A':48,'B':18,'C':12,'D':22}.items(): tracks.column_dimensions[col].width=w

baseball=wb.create_sheet('Preliminary_Calls')
baseball.append(['Preliminary Call','Count'])
for call,count in summary['baseball_calls'].items(): baseball.append([call,count])
for c in baseball[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='1F4E78')
baseball.column_dimensions['A'].width=24; baseball.column_dimensions['B'].width=12

DD=wb.create_sheet('Data_Dictionary')
DD.append(['Field','Purpose / Rule'])
rules={
 'Historical Source URL':'Original/representative SAM notice; never a downstream spending transaction alone.',
 'Open Competitive Path':'Must permit a new bidder without an existing parent BPA, IDIQ, schedule, or contract.',
 'Could GGG Compete At Time':'Historical eligibility test using actual GGG launch capabilities; no invented certifications.',
 'Mandatory-Source Status':'AbilityOne, FPI, required schedule/channel, brand or approved-source gate.',
 'Requirement Completeness':'Gate 7 remains pending until solicitation and attachments are reviewed.',
 'Payment-Method Evidence':'Unknown unless supported by the historical solicitation/award; never inferred from amount.',
 'Preliminary Supplier Exposure':'Not modeled until requirements pass and public supplier evidence exists.',
 'Preliminary Baseball Call':'Frozen only after final-20 selection and before outcome reconstruction.',
 'Award Amount':'Answer key only; not available cash and not proof of payment method or margin.',
 'Source Extract Row':'Audit pointer to the official SAM full CSV extract used for intake.'}
for h in headers: DD.append([h,rules.get(h,'Intake or analysis field defined by Historical Open Solicitation Test 01 instructions.')])
for c in DD[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='1F4E78')
DD.column_dimensions['A'].width=38; DD.column_dimensions['B'].width=100
DD.freeze_panes='A2'

meta=wb.create_sheet('Methodology')
method=[
 ['Item','Value'],['Test','Green Gregory Group — Historical Open Solicitation Test 01'],
 ['Official source','SAM.gov Contract Opportunities Full CSV extract'],
 ['Primary period','Posted 2023-01-01 through 2024-12-31; response deadline no later than 2024-12-31'],
 ['Selection unit','Distinct solicitation family; amendments and notice versions collapsed'],
 ['Selection independence','Award availability did not control sample selection'],
 ['Sampling correction','The embedded-award-only pass was rejected after producing 98/100 rejects and 88/100 DoD concentration'],
 ['Current status','100-row intake complete; 42 records require archive/attachment review before final-20 selection'],
 ['Security','No API keys, card data, bank data, nonpublic supplier data, or restricted information stored'],
]
for r in method: meta.append(r)
for c in meta[1]: c.font=Font(bold=True,color='FFFFFF'); c.fill=PatternFill('solid',fgColor='1F4E78')
meta.column_dimensions['A'].width=30; meta.column_dimensions['B'].width=110

wb.save(DEST)
print(DEST)
